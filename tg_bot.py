"""
Telegram-бот для SERP-парсингу.
Функції:
  - Ручний запуск парсингу (вибір проєкту + сторінок)
  - Керування проєктами: додати / редагувати / видалити
  - Остання статистика
"""

import asyncio, aiohttp, datetime, json, os, sys, time
import urllib.error, urllib.parse, urllib.request
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────
# НАЛАШТУВАННЯ
# ─────────────────────────────────────────
TOKEN   = os.environ.get("TG_BOT_TOKEN", "8683656869:AAEZb8YZmgjUFCHXVFQ1f_C6qq-Nx64dBKU")
CHAT_ID = os.environ.get("TG_CHAT_ID",   "909587225")

DATA_FILE    = "data/projects.json"
HISTORY_FILE = "data/history.json"

# ─────────────────────────────────────────
# СТАН ДІАЛОГУ (wizard для створення/редагування)
# ─────────────────────────────────────────
# state[chat_id] = {"step": str, "data": dict, "edit_idx": int|None}
state: dict = {}

STEPS = ["name", "location", "gl", "hl", "api_key", "keywords", "target_domains", "pages"]

STEP_PROMPTS = {
    "name":           ("✏️ <b>Назва проєкту</b>\n\nВведи назву (наприклад: <code>Germany Casino</code>):", None),
    "location":       ("🌍 <b>Локація</b>\n\nВведи локацію для Serper.dev\n(наприклад: <code>Germany</code>, <code>United States</code>):", None),
    "gl":             ("🏳️ <b>Країна (gl)</b>\n\nДвобуквений код країни для Google\n(наприклад: <code>de</code>, <code>us</code>, <code>ua</code>):", None),
    "hl":             ("🗣 <b>Мова (hl)</b>\n\nДвобуквений код мови\n(наприклад: <code>de</code>, <code>en</code>, <code>uk</code>):", None),
    "api_key":        ("🔑 <b>Serper.dev API Key</b>\n\nВстав свій API ключ з <a href='https://serper.dev'>serper.dev</a>:", None),
    "keywords":       ("📝 <b>Ключові слова</b>\n\nВведи ключові слова — кожне з <b>нового рядка</b>:\n\n<code>online casino\nbest casino germany\ncasino bonus</code>", None),
    "target_domains": ("🎯 <b>Цільові домени</b>\n\nВведи домени для відстеження — кожен з нового рядка:\n\n<code>example.com\nmysite.de</code>\n\n<i>Субдомени підтягнуться автоматично.</i>", None),
    "pages":          ("📄 <b>Кількість сторінок за замовчуванням</b>\n\nВведи число від 1 до 10\n<i>(1 стор. = 10 результатів SERP)</i>:", None),
}

STEP_LABELS = {
    "name":           "Назва",
    "location":       "Локація",
    "gl":             "gl (країна)",
    "hl":             "hl (мова)",
    "api_key":        "API Key",
    "keywords":       "Ключові слова",
    "target_domains": "Цільові домени",
    "pages":          "Сторінок за замовч.",
}

# ─────────────────────────────────────────
# TELEGRAM API
# ─────────────────────────────────────────

def tg_get(method: str, params: dict = None) -> dict:
    url = f"https://api.telegram.org/bot{TOKEN}/{method}"
    if params:
        url += "?" + urllib.parse.urlencode(params)
    try:
        with urllib.request.urlopen(url, timeout=35) as r:
            return json.loads(r.read().decode())
    except urllib.error.HTTPError as e:
        print(f"[GET] {method} {e.code}: {e.read().decode('utf-8','ignore')}")
        return {}
    except Exception as e:
        print(f"[GET] {method}: {e}")
        return {}

def tg_post(method: str, payload: dict) -> dict:
    url  = f"https://api.telegram.org/bot{TOKEN}/{method}"
    data = json.dumps(payload).encode()
    req  = urllib.request.Request(url, data=data,
                                  headers={"Content-Type": "application/json"},
                                  method="POST")
    try:
        with urllib.request.urlopen(req, timeout=15) as r:
            return json.loads(r.read().decode())
    except urllib.error.HTTPError as e:
        print(f"[POST] {method} {e.code}: {e.read().decode('utf-8','ignore')}")
        return {}
    except Exception as e:
        print(f"[POST] {method}: {e}")
        return {}

def get_updates(offset=None) -> list:
    p = {"timeout": 25, "allowed_updates": "message,callback_query"}
    if offset is not None:
        p["offset"] = offset
    return tg_get("getUpdates", p).get("result", [])

def send_msg(chat_id, text, markup=None, parse_mode="HTML", disable_preview=True):
    p = {"chat_id": chat_id, "text": text, "parse_mode": parse_mode,
         "disable_web_page_preview": disable_preview}
    if markup:
        p["reply_markup"] = markup
    return tg_post("sendMessage", p)

def edit_msg(chat_id, msg_id, text, markup=None, parse_mode="HTML"):
    p = {"chat_id": chat_id, "message_id": msg_id, "text": text,
         "parse_mode": parse_mode, "disable_web_page_preview": True}
    if markup:
        p["reply_markup"] = markup
    return tg_post("editMessageText", p)

def answer_cb(cb_id, text=""):
    tg_post("answerCallbackQuery", {"callback_query_id": cb_id, "text": text})

def send_doc(chat_id, filepath, caption=""):
    url      = f"https://api.telegram.org/bot{TOKEN}/sendDocument"
    boundary = "----BotBoundary42"
    filename = os.path.basename(filepath)
    with open(filepath, "rb") as f:
        file_data = f.read()
    def field(name, val):
        return (f"--{boundary}\r\nContent-Disposition: form-data; "
                f'name="{name}"\r\n\r\n{val}\r\n').encode()
    body = (field("chat_id", str(chat_id))
            + field("caption", caption)
            + field("parse_mode", "HTML")
            + (f"--{boundary}\r\nContent-Disposition: form-data; "
               f'name="document"; filename="{filename}"\r\n'
               f"Content-Type: application/vnd.openxmlformats-officedocument"
               f".spreadsheetml.sheet\r\n\r\n").encode()
            + file_data
            + f"\r\n--{boundary}--\r\n".encode())
    req = urllib.request.Request(
        url, data=body,
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
        method="POST")
    try:
        with urllib.request.urlopen(req, timeout=60) as r:
            return r.status == 200
    except Exception as e:
        print(f"[send_doc] {e}")
        return False

# ─────────────────────────────────────────
# ДАНІ
# ─────────────────────────────────────────

def load_projects() -> list:
    if not os.path.exists(DATA_FILE):
        return []
    try:
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            return json.load(f).get("projects", [])
    except Exception:
        return []

def save_projects(projects: list):
    os.makedirs(os.path.dirname(DATA_FILE), exist_ok=True)
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump({"projects": projects}, f, ensure_ascii=False, indent=2)

def load_history() -> list:
    if not os.path.exists(HISTORY_FILE):
        return []
    try:
        with open(HISTORY_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []

def save_history(project: dict, results: list):
    history = load_history()
    history.append({
        "timestamp":      datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "project":        project["name"],
        "location":       project.get("location", ""),
        "pages":          project.get("pages", 1),
        "target_domains": project.get("target_domains", []) or [],
        "results":        results,
    })
    if len(history) > 50:
        history = history[-50:]
    os.makedirs(os.path.dirname(HISTORY_FILE), exist_ok=True)
    with open(HISTORY_FILE, "w", encoding="utf-8") as f:
        json.dump(history, f, ensure_ascii=False, indent=2)

# ─────────────────────────────────────────
# ДОМЕНИ
# ─────────────────────────────────────────

def norm(d: str) -> str:
    d = (d or "").strip().lower()
    return d[4:] if d.startswith("www.") else d

def get_root(domain: str, targets: list):
    d = norm(domain)
    for t in targets:
        tn = norm(t)
        if tn and (d == tn or d.endswith("." + tn)):
            return tn
    return None

def enrich(results: list, targets: list) -> list:
    out = []
    for r in results:
        dc   = r.get("domain_clean") or r.get("domain", "")
        root = get_root(dc, targets)
        out.append({**r, "is_target": bool(root), "target_root": root})
    return out

# ─────────────────────────────────────────
# SERP PARSER
# ─────────────────────────────────────────

def _extract_domain(url: str):
    try:
        host = url.split("://", 1)[-1].split("/")[0].split(":")[0].lower()
        return host, host[4:] if host.startswith("www.") else host
    except Exception:
        return url, url

async def fetch_serp(session, api_key, keyword, location, gl, hl, page, sem):
    async with sem:
        headers = {"X-API-KEY": api_key, "Content-Type": "application/json"}
        payload = {"q": keyword, "location": location, "gl": gl, "hl": hl,
                   "num": 10, "page": page, "autocorrect": False}
        try:
            async with session.post("https://google.serper.dev/search",
                                    json=payload, headers=headers) as resp:
                if resp.status != 200:
                    print(f"[serp] {resp.status} '{keyword}' p{page}")
                    return []
                data = await resp.json()
                results = []
                for idx, item in enumerate(data.get("organic", []), 1):
                    link = item.get("link", "")
                    domain, domain_clean = _extract_domain(link)
                    results.append({
                        "keyword": keyword, "position": (page-1)*10+idx,
                        "domain": domain, "domain_clean": domain_clean,
                        "title": item.get("title",""), "snippet": item.get("snippet",""),
                        "url": link,
                    })
                return results
        except Exception as e:
            print(f"[serp] error '{keyword}' p{page}: {e}")
            return []

async def parse_project(proj: dict) -> list:
    api_key  = proj.get("api_key", "")
    keywords = proj.get("keywords", [])
    location = proj.get("location", "Germany")
    gl       = proj.get("gl", "de")
    hl       = proj.get("hl", "de")
    pages    = proj.get("pages", 1)
    if not api_key or not keywords:
        return []
    sem   = asyncio.Semaphore(10)
    tasks = [(kw, pg) for kw in keywords for pg in range(1, pages+1)]
    async with aiohttp.ClientSession(timeout=aiohttp.ClientTimeout(total=30)) as session:
        coros = [fetch_serp(session, api_key, kw, location, gl, hl, pg, sem) for kw, pg in tasks]
        batches = await asyncio.gather(*coros)
    results = []
    for b in batches:
        results.extend(b)
    return results

# ─────────────────────────────────────────
# EXCEL
# ─────────────────────────────────────────

BUCKET_KEYS   = ["1-3","4-10","11-20","21-30","31-40","41-50","51-100"]
BUCKET_SCORES = {"1-3":100,"4-10":60,"11-20":40,"21-30":25,"31-40":15,"41-50":8,"51-100":3}
_thin = Side(style="thin", color="CCCCCC")
_brd  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
def _fill(c): return PatternFill(start_color=c, end_color=c, fill_type="solid")
def _hdr(ws):
    ws.row_dimensions[1].height = 22
    for cell in ws[1]:
        cell.font=Font(bold=True,color="FFFFFF",size=10); cell.fill=_fill("1F4E78")
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        cell.border=_brd
def _row(ws, rn, it=False):
    for cell in ws[rn]:
        if it: cell.fill=_fill("C6EFCE")
        cell.border=_brd; cell.font=Font(size=10)
    ws.row_dimensions[rn].height=16
def bkt(pos):
    if not isinstance(pos,int): return ">100"
    for lo,hi,k in [(1,3,"1-3"),(4,10,"4-10"),(11,20,"11-20"),(21,30,"21-30"),
                    (31,40,"31-40"),(41,50,"41-50"),(51,100,"51-100")]:
        if lo<=pos<=hi: return k
    return ">100"
def score(b): return sum(b.get(k,0)*BUCKET_SCORES[k] for k in BUCKET_KEYS)

def export_excel(results, filename, targets, history):
    wb = Workbook()
    ws = wb.active; ws.title = "Results"
    ws.append(["#","Keyword","Position","Domain","Title","Snippet","URL","Target","Root"])
    _hdr(ws); ws.freeze_panes="A2"
    for i,r in enumerate(results,1):
        it=r.get("is_target",False)
        ws.append([i,r.get("keyword",""),r.get("position"),r.get("domain",""),
                   r.get("title",""),r.get("snippet",""),r.get("url",""),
                   "✅" if it else "",r.get("target_root") or ""])
        _row(ws,ws.max_row,it)
    for col,w in zip("ABCDEFGHI",[5,32,8,28,38,42,52,6,22]):
        ws.column_dimensions[col].width=w
    if history:
        last=history[-1]["results"]
        ws2=wb.create_sheet("Target Stats")
        ws2.append(["Domain","Root"]+[f"Pos {k}" for k in BUCKET_KEYS]+["Score","KW","Keywords"])
        _hdr(ws2)
        db=defaultdict(lambda:{k:0 for k in BUCKET_KEYS}); dk=defaultdict(set); dr={}
        for r in last:
            if not r.get("is_target"): continue
            pos=r.get("position")
            if not isinstance(pos,int): continue
            b2=bkt(pos)
            if b2 not in BUCKET_KEYS: continue
            d=r["domain"]; db[d][b2]+=1; dk[d].add(r["keyword"]); dr[d]=r.get("target_root","")
        for d,bkts in sorted(db.items(),key=lambda x:score(x[1]),reverse=True):
            ws2.append([d,dr.get(d,"")]+[bkts[k] for k in BUCKET_KEYS]
                       +[score(bkts),len(dk[d]),"; ".join(sorted(dk[d]))])
            _row(ws2,ws2.max_row,True)
        ws3=wb.create_sheet("Dynamics")
        labels=[e.get("timestamp","")[:10] for e in history]
        ws3.append(["Keyword","Domain","Current","Trend"]+labels+["Avg","Best","Worst"])
        _hdr(ws3)
        combos={(r["keyword"],r["domain"]) for e in history for r in e["results"] if r.get("is_target")}
        for kw,dom in sorted(combos):
            positions=[]
            for e in history:
                found=[r["position"] for r in e["results"] if r.get("is_target")
                       and r["keyword"]==kw and r["domain"]==dom and isinstance(r.get("position"),int)]
                positions.append(min(found) if found else None)
            valid=[p for p in positions if p is not None]
            if not valid: continue
            cur=valid[-1]
            diff=(valid[-2]-cur) if len(valid)>=2 else 0
            trend="New" if len(valid)==1 else (f"↑{diff}" if diff>0 else (f"↓{abs(diff)}" if diff<0 else "="))
            ws3.append([kw,dom,cur,trend]+[p if p is not None else "—" for p in positions]
                       +[round(sum(valid)/len(valid),1),min(valid),max(valid)])
            _row(ws3,ws3.max_row,True)
    wb.save(filename)

# ─────────────────────────────────────────
# ЗВІТ
# ─────────────────────────────────────────

def build_report(proj, results, duration, end_time, pages):
    dom=defaultdict(lambda:{"pos":[],"kw":set()})
    for r in results:
        if not r.get("is_target"): continue
        pos=r.get("position")
        if not isinstance(pos,int): continue
        root=r.get("target_root") or r.get("domain","")
        dom[root]["pos"].append(pos); dom[root]["kw"].add(r["keyword"])
    lines=[]
    for d in sorted(dom):
        pos=dom[d]["pos"]
        if not pos: continue
        kw=len(dom[d]["kw"]); avg=round(sum(pos)/len(pos),1)
        t=[sum(1 for p in pos if p<=3),sum(1 for p in pos if 4<=p<=10),
           sum(1 for p in pos if 11<=p<=20),sum(1 for p in pos if 21<=p<=50),
           sum(1 for p in pos if 51<=p<=100)]
        icons=["🥇 1–3","🟢 4–10","🟡 11–20","🟠 21–50","🔴 51–100"]
        bars="  ".join(f"{ic}: {n}" for ic,n in zip(icons,t) if n)
        lines.append(f"🌐 <b>{d}</b>\n   {bars}\n   📌 KW: {kw}  ⌀ поз: {avg}")
    total_kw=len({r["keyword"] for r in results})
    block="\n\n".join(lines) if lines else "— немає попадань —"
    return (f"━━━━━━━━━━━━━━━━━━━━\n📊 <b>SERP ЗВІТ</b> [ручний]\n━━━━━━━━━━━━━━━━━━━━\n"
            f"📁 <b>{proj['name']}</b>\n🗓 {end_time.strftime('%d.%m.%Y %H:%M')}\n"
            f"🌍 {proj.get('location','')}  gl={proj.get('gl','')}  hl={proj.get('hl','')}\n"
            f"⏱ {duration:.0f} сек\n━━━━━━━━━━━━━━━━━━━━\n"
            f"🔑 Ключів: {total_kw}  📄 Сторінок: {pages}\n━━━━━━━━━━━━━━━━━━━━\n"
            f"📌 <b>ПОЗИЦІЇ ПО ДОМЕНАХ</b>\n━━━━━━━━━━━━━━━━━━━━\n{block}\n━━━━━━━━━━━━━━━━━━━━")

# ─────────────────────────────────────────
# КЛАВІАТУРИ
# ─────────────────────────────────────────

def kb_main():
    return {"inline_keyboard": [
        [{"text": "🚀  Запустити парсинг",    "callback_data": "go:parse"}],
        [{"text": "📁  Керування проєктами",   "callback_data": "go:manage"}],
        [{"text": "📈  Остання статистика",    "callback_data": "go:stats"}],
    ]}

def kb_manage():
    return {"inline_keyboard": [
        [{"text": "➕  Додати проєкт",         "callback_data": "proj:new"}],
        [{"text": "✏️  Редагувати проєкт",     "callback_data": "proj:edit_list"}],
        [{"text": "🗑  Видалити проєкт",        "callback_data": "proj:del_list"}],
        [{"text": "📋  Переглянути всі",        "callback_data": "proj:view_all"}],
        [{"text": "◀️  Головне меню",           "callback_data": "go:main"}],
    ]}

def kb_parse_list(projects):
    rows = [[{"text": f"📁 {p['name']}  ({len(p.get('keywords',[]))} KW)",
              "callback_data": f"parse:proj:{i}"}] for i,p in enumerate(projects)]
    rows.append([{"text": "◀️ Назад", "callback_data": "go:main"}])
    return {"inline_keyboard": rows}

def kb_pages(pi):
    row = [{"text": f"{n} стор.", "callback_data": f"parse:pages:{pi}:{n}"} for n in [1,2,3,5,10]]
    return {"inline_keyboard": [row, [{"text": "◀️ Назад", "callback_data": "go:parse"}]]}

def kb_confirm_parse(pi, pg):
    return {"inline_keyboard": [[
        {"text": "✅ Запустити",  "callback_data": f"parse:run:{pi}:{pg}"},
        {"text": "❌ Скасувати", "callback_data": "go:main"},
    ]]}

def kb_edit_list(projects):
    rows = [[{"text": f"✏️ {p['name']}", "callback_data": f"proj:edit:{i}"}]
            for i,p in enumerate(projects)]
    rows.append([{"text": "◀️ Назад", "callback_data": "go:manage"}])
    return {"inline_keyboard": rows}

def kb_edit_fields(pi):
    rows = [[{"text": f"  {STEP_LABELS[s]}", "callback_data": f"proj:edit_field:{pi}:{s}"}]
            for s in STEPS]
    rows.append([{"text": "◀️ Назад", "callback_data": "proj:edit_list"}])
    return {"inline_keyboard": rows}

def kb_del_list(projects):
    rows = [[{"text": f"🗑 {p['name']}", "callback_data": f"proj:del_confirm:{i}"}]
            for i,p in enumerate(projects)]
    rows.append([{"text": "◀️ Назад", "callback_data": "go:manage"}])
    return {"inline_keyboard": rows}

def kb_del_confirm(pi, name):
    return {"inline_keyboard": [[
        {"text": f"✅ Так, видалити «{name}»", "callback_data": f"proj:del_do:{pi}"},
        {"text": "❌ Скасувати",               "callback_data": "go:manage"},
    ]]}

def kb_back_manage():
    return {"inline_keyboard": [[{"text": "◀️ Назад", "callback_data": "go:manage"}]]}

def kb_back_main():
    return {"inline_keyboard": [[{"text": "◀️ Головне меню", "callback_data": "go:main"}]]}

# ─────────────────────────────────────────
# ТЕКСТИ ЕКРАНІВ
# ─────────────────────────────────────────

def t_main():
    return ("╔════════════════════════╗\n"
            "║  🤖  <b>SERP Parser Bot</b>   ║\n"
            "╚════════════════════════╝\n\n"
            "Що хочеш зробити?")

def t_manage():
    projects = load_projects()
    n = len(projects)
    count = f"Проєктів: <b>{n}</b>" if n else "Проєктів ще немає."
    return f"📁 <b>Керування проєктами</b>\n\n{count}"

def t_project_card(p):
    kw  = p.get("keywords", [])
    dom = p.get("target_domains", []) or []
    return (f"📁 <b>{p['name']}</b>\n\n"
            f"🌍 Локація: <code>{p.get('location','—')}</code>\n"
            f"🏳️ gl: <code>{p.get('gl','—')}</code>  🗣 hl: <code>{p.get('hl','—')}</code>\n"
            f"📄 Сторінок за замовч.: <code>{p.get('pages',1)}</code>\n"
            f"🔑 Ключових слів: <b>{len(kw)}</b>\n"
            f"🎯 Цільових доменів: <b>{len(dom)}</b>\n"
            f"   {', '.join(dom) if dom else '—'}")

def t_stats():
    history = load_history()
    if not history:
        return "📈 <b>Статистика</b>\n\nІсторія порожня."
    lines = ["📈 <b>Остання статистика</b>\n"]
    seen  = set()
    for e in reversed(history):
        n = e.get("project","—")
        if n in seen: continue
        seen.add(n)
        ts=e.get("timestamp","")[:16]; res=e.get("results",[])
        total=len({r["keyword"] for r in res})
        targets=[r for r in res if r.get("is_target")]
        top3=sum(1 for r in targets if isinstance(r.get("position"),int) and r["position"]<=3)
        top10=sum(1 for r in targets if isinstance(r.get("position"),int) and r["position"]<=10)
        lines.append(f"📁 <b>{n}</b>  <i>({ts})</i>\n"
                     f"   🔑 {total} KW  |  🎯 {len(targets)} попадань\n"
                     f"   🥇 Топ-3: {top3}  |  🟢 Топ-10: {top10}")
    return "\n\n".join(lines)

# ─────────────────────────────────────────
# WIZARD: крок за кроком
# ─────────────────────────────────────────

def wizard_start_new(chat_id):
    state[chat_id] = {"step": "name", "data": {}, "edit_idx": None, "edit_field": None}
    send_msg(chat_id, "🆕 <b>Новий проєкт</b>\n\nБудемо заповнювати крок за кроком.\n\n"
             + STEP_PROMPTS["name"][0])

def wizard_start_edit_field(chat_id, pi, field):
    projects = load_projects()
    if pi >= len(projects):
        send_msg(chat_id, "❌ Проєкт не знайдено.", markup=kb_back_manage())
        return
    proj = projects[pi]
    state[chat_id] = {"step": field, "data": dict(proj), "edit_idx": pi, "edit_field": field}
    current = proj.get(field, "")
    if isinstance(current, list):
        current = "\n".join(current)
    send_msg(chat_id,
             f"✏️ Редагуєш: <b>{STEP_LABELS[field]}</b> проєкту «{proj['name']}»\n\n"
             f"Поточне значення:\n<code>{current}</code>\n\n"
             + STEP_PROMPTS[field][0])

def wizard_process(chat_id, text):
    """Обробляє поточний крок wizard і переходить до наступного."""
    s = state.get(chat_id)
    if not s:
        return False  # не в wizard

    step = s["step"]
    data = s["data"]

    # Парсимо введення
    if step == "pages":
        try:
            val = int(text.strip())
            if not 1 <= val <= 10:
                raise ValueError
        except ValueError:
            send_msg(chat_id, "⚠️ Введи ціле число від 1 до 10:")
            return True
        data[step] = val
    elif step in ("keywords", "target_domains"):
        lines = [l.strip() for l in text.strip().splitlines() if l.strip()]
        if not lines:
            send_msg(chat_id, "⚠️ Введи хоча б одне значення:")
            return True
        data[step] = lines
    else:
        val = text.strip()
        if not val:
            send_msg(chat_id, "⚠️ Не може бути порожнім. Спробуй ще раз:")
            return True
        data[step] = val

    # Якщо редагуємо одне поле — одразу зберігаємо
    if s["edit_field"] is not None:
        projects = load_projects()
        pi = s["edit_idx"]
        projects[pi][step] = data[step]
        save_projects(projects)
        del state[chat_id]
        send_msg(chat_id,
                 f"✅ <b>{STEP_LABELS[step]}</b> оновлено!\n\n"
                 + t_project_card(projects[pi]),
                 markup=kb_edit_fields(pi))
        return True

    # Інакше — переходимо до наступного кроку
    idx  = STEPS.index(step)
    if idx + 1 < len(STEPS):
        next_step = STEPS[idx + 1]
        s["step"] = next_step
        progress  = f"({idx+1}/{len(STEPS)})"
        send_msg(chat_id, f"{progress} {STEP_PROMPTS[next_step][0]}")
    else:
        # Всі кроки пройдено — зберігаємо
        projects = load_projects()
        projects.append(data)
        save_projects(projects)
        del state[chat_id]
        send_msg(chat_id,
                 f"✅ <b>Проєкт «{data['name']}» створено!</b>\n\n"
                 + t_project_card(data),
                 markup=kb_manage())
    return True

# ─────────────────────────────────────────
# ЗАПУСК ПАРСИНГУ
# ─────────────────────────────────────────

def run_parsing(chat_id, msg_id, proj: dict, pages: int):
    proj_run = dict(proj); proj_run["pages"] = pages
    kw = len(proj_run.get("keywords", []))
    edit_msg(chat_id, msg_id,
             f"⏳ <b>Парсинг запущено...</b>\n\n"
             f"📁 {proj_run['name']}\n"
             f"🔑 {kw} ключів  ×  📄 {pages} стор. = ~{kw*pages} запитів\n\n"
             f"<i>Зачекайте...</i>")
    t0      = datetime.datetime.now()
    results = asyncio.run(parse_project(proj_run))
    t1      = datetime.datetime.now()
    dur     = (t1 - t0).total_seconds()
    targets = proj_run.get("target_domains", []) or []
    results = enrich(results, targets)
    save_history(proj_run, results)
    hist    = [h for h in load_history() if h.get("project") == proj_run["name"]]
    fname   = f"SERP_{proj_run['name']}_{t1.strftime('%Y%m%d_%H%M')}.xlsx"
    export_excel(results, fname, targets, hist)
    report  = build_report(proj_run, results, dur, t1, pages)
    edit_msg(chat_id, msg_id, f"✅ <b>Готово!</b>\n\n{report}\n\n📎 Відправляю файл...")
    if not send_doc(chat_id, fname,
                    caption=f"📊 {proj_run['name']} | {t1.strftime('%d.%m.%Y %H:%M')}"):
        send_msg(chat_id, "⚠️ Не вдалося відправити файл.")
    try: os.remove(fname)
    except Exception: pass
    send_msg(chat_id, t_main(), markup=kb_main())

# ─────────────────────────────────────────
# ОБРОБКА CALLBACK
# ─────────────────────────────────────────

def on_callback(cq):
    cq_id   = cq["id"]
    chat_id = cq["message"]["chat"]["id"]
    msg_id  = cq["message"]["message_id"]
    from_id = str(cq["from"]["id"])
    data    = cq.get("data","")

    answer_cb(cq_id)

    if CHAT_ID and from_id != CHAT_ID:
        print(f"[auth] denied from_id={from_id}")
        return

    # Скидаємо wizard якщо користувач натиснув кнопку
    if chat_id in state:
        del state[chat_id]

    projects = load_projects()

    # ── НАВІГАЦІЯ ──
    if data == "go:main":
        edit_msg(chat_id, msg_id, t_main(), markup=kb_main())

    elif data == "go:parse":
        if not projects:
            edit_msg(chat_id, msg_id,
                     "📭 Проєктів немає.\n\nСпочатку створи проєкт через «Керування проєктами».",
                     markup=kb_back_main())
        else:
            edit_msg(chat_id, msg_id, "🚀 <b>Оберіть проєкт для парсингу:</b>",
                     markup=kb_parse_list(projects))

    elif data == "go:manage":
        edit_msg(chat_id, msg_id, t_manage(), markup=kb_manage())

    elif data == "go:stats":
        edit_msg(chat_id, msg_id, t_stats(), markup=kb_back_main())

    # ── ПАРСИНГ ──
    elif data.startswith("parse:proj:"):
        pi = int(data.split(":")[2])
        if pi >= len(projects):
            edit_msg(chat_id, msg_id, "❌ Проєкт не знайдено.")
            return
        p  = projects[pi]
        kw = len(p.get("keywords", []))
        edit_msg(chat_id, msg_id,
                 f"📁 <b>{p['name']}</b>\n🌍 {p.get('location','—')}  |  🔑 {kw} KW\n\n"
                 f"📄 <b>Скільки сторінок парсити?</b>\n<i>1 стор. = 10 результатів SERP</i>",
                 markup=kb_pages(pi))

    elif data.startswith("parse:pages:"):
        _, _, pi, pg = data.split(":")
        pi, pg = int(pi), int(pg)
        if pi >= len(projects):
            edit_msg(chat_id, msg_id, "❌ Проєкт не знайдено.")
            return
        p  = projects[pi]
        kw = len(p.get("keywords", []))
        edit_msg(chat_id, msg_id,
                 f"🔎 <b>Підтвердження</b>\n\n📁 {p['name']}\n🌍 {p.get('location','—')}\n"
                 f"🔑 {kw} ключів  ×  📄 {pg} стор. = ~{kw*pg} запитів\n\nЗапустити?",
                 markup=kb_confirm_parse(pi, pg))

    elif data.startswith("parse:run:"):
        _, _, pi, pg = data.split(":")
        pi, pg = int(pi), int(pg)
        if pi >= len(projects):
            edit_msg(chat_id, msg_id, "❌ Проєкт не знайдено.")
            return
        run_parsing(chat_id, msg_id, projects[pi], pg)

    # ── КЕРУВАННЯ ПРОЄКТАМИ ──
    elif data == "proj:new":
        edit_msg(chat_id, msg_id, "⬇️ Починаємо створення нового проєкту...")
        wizard_start_new(chat_id)

    elif data == "proj:edit_list":
        if not projects:
            edit_msg(chat_id, msg_id, "📭 Немає проєктів для редагування.", markup=kb_back_manage())
        else:
            edit_msg(chat_id, msg_id, "✏️ <b>Який проєкт редагуємо?</b>",
                     markup=kb_edit_list(projects))

    elif data.startswith("proj:edit:"):
        pi = int(data.split(":")[2])
        if pi >= len(projects):
            edit_msg(chat_id, msg_id, "❌ Проєкт не знайдено.")
            return
        edit_msg(chat_id, msg_id,
                 f"✏️ <b>Редагування: «{projects[pi]['name']}»</b>\n\n"
                 "Яке поле змінити?",
                 markup=kb_edit_fields(pi))

    elif data.startswith("proj:edit_field:"):
        parts = data.split(":")
        pi, field = int(parts[2]), parts[3]
        edit_msg(chat_id, msg_id, "⬇️ Введи нове значення нижче...")
        wizard_start_edit_field(chat_id, pi, field)

    elif data == "proj:del_list":
        if not projects:
            edit_msg(chat_id, msg_id, "📭 Немає проєктів.", markup=kb_back_manage())
        else:
            edit_msg(chat_id, msg_id, "🗑 <b>Який проєкт видалити?</b>",
                     markup=kb_del_list(projects))

    elif data.startswith("proj:del_confirm:"):
        pi = int(data.split(":")[2])
        if pi >= len(projects):
            edit_msg(chat_id, msg_id, "❌ Проєкт не знайдено.")
            return
        edit_msg(chat_id, msg_id,
                 f"⚠️ Видалити проєкт <b>«{projects[pi]['name']}»</b>?\n\nЦю дію не можна скасувати.",
                 markup=kb_del_confirm(pi, projects[pi]['name']))

    elif data.startswith("proj:del_do:"):
        pi = int(data.split(":")[2])
        if pi >= len(projects):
            edit_msg(chat_id, msg_id, "❌ Проєкт не знайдено.")
            return
        name = projects[pi]["name"]
        projects.pop(pi)
        save_projects(projects)
        edit_msg(chat_id, msg_id,
                 f"🗑 Проєкт <b>«{name}»</b> видалено.\n\n{t_manage()}",
                 markup=kb_manage())

    elif data == "proj:view_all":
        if not projects:
            edit_msg(chat_id, msg_id, "📭 Проєктів немає.", markup=kb_back_manage())
        else:
            cards = "\n\n────────────────────\n\n".join(t_project_card(p) for p in projects)
            edit_msg(chat_id, msg_id, f"📋 <b>Всі проєкти ({len(projects)}):</b>\n\n{cards}",
                     markup=kb_back_manage())

# ─────────────────────────────────────────
# ОБРОБКА ПОВІДОМЛЕНЬ
# ─────────────────────────────────────────

def on_message(msg):
    chat_id = msg["chat"]["id"]
    from_id = str(msg["from"]["id"])
    text    = msg.get("text", "").strip()

    if CHAT_ID and from_id != CHAT_ID:
        print(f"[auth] denied from_id={from_id}")
        return

    # Якщо активний wizard — передаємо текст у нього
    if chat_id in state:
        wizard_process(chat_id, text)
        return

    # Інакше — головне меню
    send_msg(chat_id, t_main(), markup=kb_main())

# ─────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────

def main():
    print(f"[bot] TOKEN={TOKEN[:20]}...")
    print(f"[bot] CHAT_ID={CHAT_ID}")

    me = tg_get("getMe")
    if not me.get("ok"):
        print(f"[bot] ❌ getMe failed: {me}")
        sys.exit(1)
    print(f"[bot] ✅ @{me['result'].get('username','?')}")

    # Скидаємо старі updates
    old = tg_get("getUpdates", {"offset": -1})
    if old.get("result"):
        skip = old["result"][-1]["update_id"] + 1
        tg_get("getUpdates", {"offset": skip})
        print(f"[bot] Скинуто старі updates, offset={skip}")

    # Стартове повідомлення
    r = send_msg(CHAT_ID, t_main(), markup=kb_main())
    print(f"[bot] Старт: {'ok' if r.get('ok') else r}")
    print("[bot] Слухаю... Ctrl+C для зупинки.")

    offset = None
    while True:
        try:
            updates = get_updates(offset)
            for upd in updates:
                offset = upd["update_id"] + 1
                print(f"[bot] upd={upd['update_id']} {list(upd.keys())}")
                if "callback_query" in upd:
                    on_callback(upd["callback_query"])
                elif "message" in upd:
                    on_message(upd["message"])
        except KeyboardInterrupt:
            print("[bot] Зупинено.")
            break
        except Exception as e:
            print(f"[bot] ❌ {e}")
            time.sleep(5)

if __name__ == "__main__":
    main()
