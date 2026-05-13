from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ==========================
# СТИЛІ — мінімальні
# ==========================

TARGET_GREEN = "C6EFCE"   # зелений фон — наші домени
HEADER_BG    = "1F4E78"   # темно-синій — заголовки
HEADER_FG    = "FFFFFF"

_thin   = Side(style="thin", color="CCCCCC")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _apply_header(ws, row_num: int = 1, height: int = 22):
    ws.row_dimensions[row_num].height = height
    for cell in ws[row_num]:
        cell.font      = Font(bold=True, color=HEADER_FG, size=10)
        cell.fill      = _fill(HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _border


def _apply_data_row(ws, row_num: int, is_target: bool = False):
    bg = TARGET_GREEN if is_target else None
    for cell in ws[row_num]:
        if bg:
            cell.fill = _fill(bg)
        cell.border = _border
        cell.font   = Font(size=10)
        if not cell.alignment or cell.alignment.horizontal == "general":
            cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row_num].height = 16


def _set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


# ==========================
# ПОЗИЦІЙНІ БАКЕТИ
# ==========================

BUCKET_KEYS = ["1-3", "4-10", "11-20", "21-30", "31-40", "41-50", "51-100"]

BUCKET_SCORES = {
    "1-3":   100,
    "4-10":   60,
    "11-20":  40,
    "21-30":  25,
    "31-40":  15,
    "41-50":   8,
    "51-100":  3,
}


def bucket_for_position(pos) -> str:
    if not isinstance(pos, int):
        return ">100"
    if 1  <= pos <= 3:   return "1-3"
    if 4  <= pos <= 10:  return "4-10"
    if 11 <= pos <= 20:  return "11-20"
    if 21 <= pos <= 30:  return "21-30"
    if 31 <= pos <= 40:  return "31-40"
    if 41 <= pos <= 50:  return "41-50"
    if 51 <= pos <= 100: return "51-100"
    return ">100"


def calculate_score(buckets: dict) -> int:
    return sum(buckets.get(k, 0) * BUCKET_SCORES[k] for k in BUCKET_KEYS)


# ==========================
# ГОЛОВНА ФУНКЦІЯ
# ==========================

def export_to_excel(current_results, filename, target_domains=None, history=None):
    if target_domains is None:
        target_domains = []

    wb = Workbook()

    # Fallback — без history
    if not history:
        ws = wb.active
        ws.title = "Results"
        _build_results_sheet(ws, current_results, target_domains)
        wb.save(filename)
        return filename

    last_results = history[-1]["results"]

    ws_res = wb.active
    ws_res.title = "Results"
    _build_results_sheet(ws_res, last_results, target_domains, with_target_root=True)

    ws_target = wb.create_sheet("Target Domains Stats")
    _build_target_stats_sheet(ws_target, last_results)

    ws_pos = wb.create_sheet("Position Buckets")
    _build_position_buckets_sheet(ws_pos, last_results)

    ws_dyn = wb.create_sheet("Dynamics")
    _build_dynamics_sheet(ws_dyn, history)

    ws_hist = wb.create_sheet("History Summary")
    _build_history_summary_sheet(ws_hist, history)

    wb.save(filename)
    return filename


# ==========================
# ВКЛАДКИ
# ==========================

def _build_results_sheet(ws, results, target_domains, with_target_root: bool = False):
    target_set = set(target_domains)

    if with_target_root:
        headers = ["#", "Keyword", "Position", "Domain", "Title", "Snippet", "URL", "Target", "Root Domain"]
    else:
        headers = ["#", "Keyword", "Position", "Domain", "Title", "Snippet", "URL", "Target"]

    ws.append(headers)
    _apply_header(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row_num, r in enumerate(results, start=1):
        is_target = r.get("is_target", r.get("domain", "") in target_set)
        pos       = r.get("position")

        row_data = [
            row_num,
            r.get("keyword", ""),
            pos,
            r.get("domain", ""),
            r.get("title", ""),
            r.get("snippet", ""),
            r.get("url", ""),
            "✅" if is_target else "",
        ]
        if with_target_root:
            row_data.append(r.get("target_root") or "")

        ws.append(row_data)
        _apply_data_row(ws, ws.max_row, is_target=is_target)

        for col_i in (1, 3, 8):
            ws.cell(row=ws.max_row, column=col_i).alignment = Alignment(horizontal="center", vertical="center")

    _set_col_widths(ws, {
        "A": 5, "B": 32, "C": 10, "D": 28,
        "E": 40, "F": 45, "G": 55, "H": 8,
    })
    if with_target_root:
        ws.column_dimensions["I"].width = 25


def _build_target_stats_sheet(ws, results):
    headers = (
        ["Домен", "Root Domain", "Всього"]
        + [f"Pos {k}" for k in BUCKET_KEYS]
        + ["Score", "Кількість KW", "Ключові слова"]
    )
    ws.append(headers)
    _apply_header(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    domain_buckets  = defaultdict(lambda: {k: 0 for k in BUCKET_KEYS})
    domain_keywords = defaultdict(set)
    domain_roots    = {}

    for r in results:
        if not r.get("is_target"):
            continue
        dom = r["domain"]
        pos = r.get("position")
        if not isinstance(pos, int):
            continue
        bucket = bucket_for_position(pos)
        if bucket not in BUCKET_KEYS:
            continue
        domain_buckets[dom][bucket] += 1
        domain_keywords[dom].add(r["keyword"])
        domain_roots[dom] = r.get("target_root") or ""

    rows = []
    for dom, buckets in domain_buckets.items():
        total = sum(buckets.values())
        if total == 0:
            continue
        rows.append({
            "Domain":   dom,
            "Root":     domain_roots.get(dom, ""),
            "Total":    total,
            "Score":    calculate_score(buckets),
            "KW_count": len(domain_keywords[dom]),
            "Keywords": "; ".join(sorted(domain_keywords[dom])),
            **{f"Pos {k}": buckets[k] for k in BUCKET_KEYS},
        })
    rows.sort(key=lambda x: x["Score"], reverse=True)

    for row in rows:
        data = (
            [row["Domain"], row["Root"], row["Total"]]
            + [row[f"Pos {k}"] for k in BUCKET_KEYS]
            + [row["Score"], row["KW_count"], row["Keywords"]]
        )
        ws.append(data)
        _apply_data_row(ws, ws.max_row, is_target=True)
        for col_i in range(3, 3 + len(BUCKET_KEYS) + 3):
            ws.cell(row=ws.max_row, column=col_i).alignment = Alignment(horizontal="center", vertical="center")

    _set_col_widths(ws, {"A": 28, "B": 25})
    for i in range(3, 3 + len(BUCKET_KEYS) + 3):
        ws.column_dimensions[get_column_letter(i)].width = 12
    ws.column_dimensions[get_column_letter(3 + len(BUCKET_KEYS) + 3)].width = 60


def _build_position_buckets_sheet(ws, results):
    headers = (
        ["Домен", "Всього"]
        + [f"Pos {k}" for k in BUCKET_KEYS]
        + ["Score", "Наш?"]
    )
    ws.append(headers)
    _apply_header(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    all_domain_buckets = defaultdict(lambda: {k: 0 for k in BUCKET_KEYS})
    is_target_flag     = defaultdict(bool)

    for r in results:
        pos = r.get("position")
        if not isinstance(pos, int):
            continue
        bucket = bucket_for_position(pos)
        if bucket not in BUCKET_KEYS:
            continue
        dom = r["domain"]
        all_domain_buckets[dom][bucket] += 1
        if r.get("is_target"):
            is_target_flag[dom] = True

    rows = []
    for dom, buckets in all_domain_buckets.items():
        total = sum(buckets.values())
        if total == 0:
            continue
        rows.append({
            "Domain":    dom,
            "Total":     total,
            "Score":     calculate_score(buckets),
            "Is_Target": is_target_flag[dom],
            **{f"Pos {k}": buckets[k] for k in BUCKET_KEYS},
        })
    rows.sort(key=lambda x: x["Score"], reverse=True)

    for row in rows:
        data = (
            [row["Domain"], row["Total"]]
            + [row[f"Pos {k}"] for k in BUCKET_KEYS]
            + [row["Score"], "✅" if row["Is_Target"] else ""]
        )
        ws.append(data)
        _apply_data_row(ws, ws.max_row, is_target=row["Is_Target"])
        for col_i in range(2, 2 + len(BUCKET_KEYS) + 3):
            ws.cell(row=ws.max_row, column=col_i).alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 30
    for i in range(2, 2 + len(BUCKET_KEYS) + 3):
        ws.column_dimensions[get_column_letter(i)].width = 12


def _build_dynamics_sheet(ws, history):
    num_parses = len(history)

    parse_labels = []
    for i, entry in enumerate(history):
        ts = entry.get("timestamp", "")
        parse_labels.append(ts[:10] if len(ts) >= 10 else f"Parse {i+1}")

    headers = (
        ["Keyword", "Домен", "Поточна", "Тренд"]
        + parse_labels
        + ["Середня", "Краща", "Гірша", "URL", "Title", "Snippet"]
    )
    ws.append(headers)
    _apply_header(ws)
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = ws.dimensions

    combos = set()
    for entry in history:
        for r in entry["results"]:
            if r.get("is_target"):
                combos.add((r["keyword"], r["domain"]))

    for kw, dom in sorted(combos):
        positions = []
        for entry in history:
            found = [
                rr["position"]
                for rr in entry["results"]
                if rr.get("is_target")
                and rr["keyword"] == kw
                and rr["domain"] == dom
                and isinstance(rr.get("position"), int)
            ]
            positions.append(min(found) if found else None)

        valid = [p for p in positions if p is not None]
        if not valid:
            current, trend, avg_pos, best_pos, worst_pos = None, "—", None, None, None
        else:
            current = valid[-1]
            if len(valid) == 1:
                trend = "New"
            else:
                diff = valid[-2] - current
                trend = f"↑ {diff}" if diff > 0 else (f"↓ {abs(diff)}" if diff < 0 else "=")
            avg_pos   = round(sum(valid) / len(valid), 1)
            best_pos  = min(valid)
            worst_pos = max(valid)

        last_url = last_title = last_snippet = ""
        for rr in history[-1]["results"]:
            if rr.get("is_target") and rr["keyword"] == kw and rr["domain"] == dom:
                last_url     = rr.get("url", "")
                last_title   = rr.get("title", "")
                last_snippet = rr.get("snippet", "")
                break

        row = [kw, dom, current if current is not None else "—", trend]
        row += [p if p is not None else "—" for p in positions]
        row += [
            avg_pos   if avg_pos   is not None else "—",
            best_pos  if best_pos  is not None else "—",
            worst_pos if worst_pos is not None else "—",
            last_url, last_title, last_snippet,
        ]
        ws.append(row)
        _apply_data_row(ws, ws.max_row, is_target=True)
        for col_i in range(3, 5 + num_parses + 3):
            ws.cell(row=ws.max_row, column=col_i).alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 28
    for i in range(3, 5 + num_parses):
        ws.column_dimensions[get_column_letter(i)].width = 12
    last_col = 5 + num_parses
    ws.column_dimensions[get_column_letter(last_col)].width     = 55
    ws.column_dimensions[get_column_letter(last_col + 1)].width = 40
    ws.column_dimensions[get_column_letter(last_col + 2)].width = 55


def _build_history_summary_sheet(ws, history):
    headers = [
        "Дата", "Проєкт", "Домен",
        "Знайдено", "Середня поз.",
        "Топ 3", "Топ 10", "Топ 20", "Топ 30", "Топ 50", "Топ 100",
    ]
    ws.append(headers)
    _apply_header(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for entry in history:
        ts        = entry.get("timestamp", "")
        proj_name = entry.get("project", "")

        domain_positions = defaultdict(list)
        for r in entry.get("results", []):
            if not r.get("is_target"):
                continue
            pos = r.get("position")
            if isinstance(pos, int):
                domain_positions[r["domain"]].append(pos)

        for dom, positions in sorted(domain_positions.items()):
            if not positions:
                continue
            avg_pos = round(sum(positions) / len(positions), 1)
            ws.append([
                ts[:10] if len(ts) >= 10 else ts,
                proj_name, dom,
                len(positions), avg_pos,
                sum(1 for p in positions if p <= 3),
                sum(1 for p in positions if p <= 10),
                sum(1 for p in positions if p <= 20),
                sum(1 for p in positions if p <= 30),
                sum(1 for p in positions if p <= 50),
                sum(1 for p in positions if p <= 100),
            ])
            _apply_data_row(ws, ws.max_row, is_target=True)
            for col_i in range(4, 12):
                ws.cell(row=ws.max_row, column=col_i).alignment = Alignment(horizontal="center", vertical="center")

    _set_col_widths(ws, {"A": 13, "B": 22, "C": 28})
    for i in range(4, 12):
        ws.column_dimensions[get_column_letter(i)].width = 13